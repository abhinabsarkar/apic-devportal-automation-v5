# Import OS module
import os
# Import openpyxl to process an excel file
from openpyxl import load_workbook

def read_excel(excel_file, sheet_name):
    # Give the location of the file 
    file_path = os.path.join(os.getcwd(), excel_file) 
    # To open Workbook 
    wb = load_workbook(file_path, data_only=True)
    # Get the sheet 
    sheet = wb[sheet_name]
    # get max row count
    max_row=sheet.max_row    
    # get max column count
    max_column=sheet.max_column
    # iterate over all rows and add to list
    consumer_list = []
    for counter in range(2,max_row):     
        dev_org_name=sheet.cell(row=counter,column=1).value
        app_name=sheet.cell(row=counter,column=2).value
        plan_name=sheet.cell(row=counter,column=3).value
        # Comparing it with string No Data since the value is copied from another cell
        # If directly read from the formula cell, then it will compared with None
        if app_name != "No Data":   
            row = (dev_org_name, app_name, plan_name)
            consumer_list.append(row)
    return consumer_list